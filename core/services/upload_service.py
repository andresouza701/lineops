from __future__ import annotations

import csv
import logging
from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path

from allocations.models import LineAllocation
from core.exceptions.domain_exceptions import BusinessRuleException
from core.services.allocation_service import AllocationService
from django.db import IntegrityError, transaction
from django.utils.text import slugify

from employees.models import Employee
from telecom.models import PhoneLine, SIMcard

ALLOWED_ORIGEM_VALUES = {v.lower(): v for v in PhoneLine.Origem.values}

logger = logging.getLogger(__name__)
MIN_PHONE_NUMBER_DIGITS = 10


@dataclass
class UploadSummary:
    rows_processed: int = 0
    employees_created: int = 0
    employees_updated: int = 0
    simcards_created: int = 0
    simcards_updated: int = 0
    allocations_created: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)

    def to_dict(self) -> dict[str, int | list[str]]:
        return {
            "rows_processed": self.rows_processed,
            "employees_created": self.employees_created,
            "employees_updated": self.employees_updated,
            "simcards_created": self.simcards_created,
            "simcards_updated": self.simcards_updated,
            "allocations_created": self.allocations_created,
            "errors": self.errors,
        }


ALLOWED_EMPLOYEE_STATUSES = {value.lower(): value for value in Employee.Status.values}
ALLOWED_PHONE_LINE_STATUSES = {
    value.lower(): value for value in PhoneLine.Status.values
}

EMPLOYEE_STATUS_ALIASES = {
    "ativo": Employee.Status.ACTIVE,
    "inativo": Employee.Status.INACTIVE,
}

PHONE_LINE_STATUS_ALIASES = {
    "disponivel": PhoneLine.Status.AVAILABLE,
    "alocado": PhoneLine.Status.ALLOCATED,
    "quarentena": PhoneLine.Status.SUSPENDED,
    "cancelado": PhoneLine.Status.CANCELLED,
    "aquecendo": PhoneLine.Status.AQUECENDO,
    "novo": PhoneLine.Status.NOVO,
}


def process_upload_file(file_path: Path) -> UploadSummary:
    extension = file_path.suffix.lower()
    if extension == ".csv":
        rows = _parse_csv(file_path)
    elif extension == ".xlsx":
        rows = _parse_xlsx(file_path)
    else:
        raise ValueError("Formato de arquivo não suportado: use CSV ou XLSX.")

    return _ingest_rows(rows)


def _parse_csv(file_path: Path) -> list[dict[str, str]]:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with file_path.open("r", encoding=encoding, newline="") as csv_file:
                sample = csv_file.read(2048)
                csv_file.seek(0)

                delimiter = ","
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;")
                    delimiter = dialect.delimiter
                except csv.Error:
                    if sample.count(";") > sample.count(","):
                        delimiter = ";"

                reader = csv.DictReader(csv_file, delimiter=delimiter)
                return [_normalize_row(row) for row in reader]
        except UnicodeDecodeError:
            continue

    raise ValueError(
        "Nao foi possivel ler o CSV. Salve o arquivo em UTF-8, Windows-1252 ou Latin-1."
    )


def _parse_xlsx(file_path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ValueError(
            "Processamento XLSX indisponível: instale a dependência openpyxl."
        ) from exc

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(cell or "").strip().lower() for cell in next(rows)]
    except StopIteration:  # empty file
        return []

    normalized_rows = []
    for row in rows:
        values = {
            headers[idx]: (row[idx] if idx < len(row) else "")
            for idx in range(len(headers))
        }
        normalized_rows.append(_normalize_row(values))
    return normalized_rows


def _normalize_row(row: dict[str, object]) -> dict[str, str]:
    return {
        key.strip().lower(): _stringify(value)
        for key, value in row.items()
        if key is not None
    }


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _ingest_rows(rows: Iterable[dict[str, str]]) -> UploadSummary:
    summary = UploadSummary()

    for index, raw in enumerate(rows, start=2):
        if not any(raw.values()):
            continue

        try:
            # Isola cada linha em uma transação para evitar persistência parcial.
            with transaction.atomic():
                kind = raw.get("type", "").lower()
                if kind == "employee":
                    _upsert_employee(raw, summary)
                elif kind == "simcard":
                    _upsert_simcard(raw, summary)
                else:
                    raise ValueError("Coluna 'type' deve ser 'employee' ou 'simcard'.")
            summary.rows_processed += 1
        except (ValueError, IntegrityError) as exc:
            summary.errors.append(f"Linha {index}: {exc}")
        except Exception:
            logger.exception(
                "Unexpected failure while processing upload row",
                extra={"row_index": index},
            )
            raise

    return summary


def _upsert_employee(row: dict[str, str], summary: UploadSummary) -> None:
    required = ["full_name", "employee_id"]
    _ensure_required(row, required)
    teams = row.get("teams") or row.get("team") or row.get("department")
    if not teams:
        raise ValueError("Coluna obrigatória ausente ou vazia: teams.")

    status = _normalize_employee_status(row.get("status"))
    full_name = row["full_name"]
    employee_id = row["employee_id"]

    # full_name is the unique key (unique constraint on the model, case-insensitive).
    # employee_id (Carteira) is NOT unique — multiple employees can share the same
    # portfolio, so it must NOT be used as the lookup key.
    existing = Employee.all_objects.filter(
        full_name__iexact=full_name, is_deleted=False
    ).first()

    fields: dict[str, object] = {
        "full_name": full_name,
        "employee_id": employee_id,
        "teams": teams,
        "status": status,
        "is_deleted": False,
    }
    supervisor_email = row.get("corporate_email") or ""
    if supervisor_email:
        fields["corporate_email"] = supervisor_email

    manager_email = row.get("manager_email") or ""
    if manager_email:
        fields["manager_email"] = manager_email

    pa = row.get("pa") or ""
    if pa:
        fields["pa"] = pa

    if existing:
        for field_name, value in fields.items():
            setattr(existing, field_name, value)
        existing.save()
        summary.employees_updated += 1
    else:
        Employee.objects.create(**fields)
        summary.employees_created += 1


def _upsert_simcard(row: dict[str, str], summary: UploadSummary) -> None:
    row = _normalize_legacy_simcard_row(row)
    required = ["iccid", "carrier"]
    _ensure_required(row, required)

    phone_number = row.get("phone_number") or ""
    iccid = row["iccid"]
    line_status = _normalize_phone_line_status(row.get("status"))
    allocation_employee = _resolve_upload_allocation_employee(row, line_status)

    if allocation_employee and not phone_number:
        raise ValueError(
            "Linha vinculada por upload exige phone_number na linha de simcard."
        )

    sim_defaults = {
        "carrier": row["carrier"],
        "status": _map_phone_line_status_to_sim_status(line_status),
        "is_deleted": False,
    }

    if phone_number:
        # Primary key is the phone number: find the SIMcard through its line.
        # This allows multiple rows with the same ICCID (e.g. VIRTUAL) to each
        # produce a distinct SIM + line pair.
        existing_line = PhoneLine.all_objects.filter(phone_number=phone_number).first()
        active_allocation = (
            _get_active_allocation(existing_line)
            if existing_line is not None
            else None
        )
        if existing_line:
            simcard = existing_line.sim_card
            simcard.iccid = iccid
            for field_name, value in sim_defaults.items():
                setattr(simcard, field_name, value)
            simcard.save(update_fields=["iccid", *sim_defaults.keys(), "updated_at"])
            summary.simcards_updated += 1
        else:
            # No existing line for this number → always create a fresh SIMcard
            # so it can own its own 1-to-1 PhoneLine.
            simcard = SIMcard.objects.create(iccid=iccid, **sim_defaults)
            summary.simcards_created += 1

        origem = _normalize_origem(row.get("origem"))
        phone_line = existing_line or PhoneLine(phone_number=phone_number)
        phone_line.phone_number = phone_number
        phone_line.sim_card = simcard
        phone_line.status = _resolve_phone_line_status_for_upload(
            line_status=line_status,
            allocation_employee=allocation_employee,
            active_allocation=active_allocation,
        )
        phone_line.is_deleted = False
        if origem:
            phone_line.origem = origem
        phone_line.save()

        if allocation_employee and _sync_phone_line_allocation(
            phone_line=phone_line,
            employee=allocation_employee,
            active_allocation=active_allocation,
        ):
            summary.allocations_created += 1
    else:
        # No phone number: ICCID is the primary key.
        simcard = SIMcard.all_objects.filter(iccid=iccid).order_by("-id").first()
        if simcard is None:
            simcard = SIMcard.objects.create(iccid=iccid, **sim_defaults)
            summary.simcards_created += 1
        else:
            for field_name, value in sim_defaults.items():
                setattr(simcard, field_name, value)
            simcard.save(update_fields=[*sim_defaults.keys(), "updated_at"])
            summary.simcards_updated += 1


def _normalize_legacy_simcard_row(row: dict[str, str]) -> dict[str, str]:
    """Support older CSV layouts that omitted pa/phone_number/origem columns."""
    normalized = dict(row)

    if _has_shifted_semicolon_legacy_shape(row):
        normalized["status"] = row.get("pa", "")
        normalized["iccid"] = row.get("status", "")
        normalized["carrier"] = row.get("iccid", "")
        normalized["phone_number"] = row.get("carrier", "")
        normalized["pa"] = ""
        return normalized

    if not _has_compact_legacy_shape(row):
        return row

    normalized["status"] = row.get("teams", "")
    normalized["iccid"] = row.get("status", "")
    normalized["carrier"] = row.get("iccid", "")
    normalized["teams"] = ""
    return normalized


def _has_compact_legacy_shape(row: dict[str, str]) -> bool:
    phone_number = row.get("phone_number") or ""
    origem = row.get("origem") or ""
    teams_value = row.get("teams") or ""
    status_value = row.get("status") or ""
    iccid_value = row.get("iccid") or ""
    carrier_value = row.get("carrier") or ""
    return (
        not phone_number
        and not origem
        and teams_value
        and status_value
        and iccid_value
        and not carrier_value
    )


def _has_shifted_semicolon_legacy_shape(row: dict[str, str]) -> bool:
    return (
        not (row.get("phone_number") or "")
        and not (row.get("origem") or "")
        and not (row.get("teams") or "")
        and (row.get("pa") or "")
        and (row.get("status") or "")
        and (row.get("iccid") or "")
        and _looks_like_phone_number(row.get("carrier") or "")
    )


def _looks_like_phone_number(value: str) -> bool:
    digits = "".join(char for char in value if char.isdigit())
    return len(digits) >= MIN_PHONE_NUMBER_DIGITS


def _resolve_upload_allocation_employee(
    row: dict[str, str], line_status: str
) -> Employee | None:
    employee_name = (row.get("full_name") or "").strip()
    if not employee_name:
        return None

    if line_status != PhoneLine.Status.ALLOCATED:
        raise ValueError(
            "Linha vinculada por upload deve usar status ALLOCATED na linha de simcard."
        )

    employee = Employee.objects.filter(
        full_name__iexact=employee_name,
        is_deleted=False,
        status=Employee.Status.ACTIVE,
    ).first()
    if employee is None:
        raise ValueError(
            f"Usuario ativo nao encontrado para vinculacao da linha: {employee_name}."
        )
    return employee


def _get_active_allocation(phone_line: PhoneLine | None) -> LineAllocation | None:
    if phone_line is None:
        return None
    return (
        LineAllocation.objects.select_related("employee")
        .filter(phone_line=phone_line, is_active=True)
        .first()
    )


def _resolve_phone_line_status_for_upload(
    *,
    line_status: str,
    allocation_employee: Employee | None,
    active_allocation: LineAllocation | None,
) -> str:
    if allocation_employee is None:
        return line_status

    if (
        active_allocation is not None
        and active_allocation.employee_id == allocation_employee.id
    ):
        return PhoneLine.Status.ALLOCATED

    return PhoneLine.Status.AVAILABLE


def _sync_phone_line_allocation(
    *,
    phone_line: PhoneLine,
    employee: Employee,
    active_allocation: LineAllocation | None,
) -> bool:
    if active_allocation is None:
        active_allocation = _get_active_allocation(phone_line)

    if active_allocation and active_allocation.employee_id == employee.id:
        return False

    if active_allocation is not None:
        AllocationService.release_line(active_allocation, released_by=None)

    try:
        AllocationService.allocate_line(
            employee=employee,
            phone_line=phone_line,
            allocated_by=None,
        )
    except BusinessRuleException as exc:
        raise ValueError(str(exc)) from exc

    return True


def _ensure_required(row: dict[str, str], required_fields: list[str]) -> None:
    missing = [field for field in required_fields if not row.get(field)]
    if missing:
        joined = ", ".join(missing)
        raise ValueError(f"Colunas obrigatórias ausentes ou vazias: {joined}.")


def _normalize_employee_status(raw_status: str | None) -> str:
    if not raw_status:
        return Employee.Status.ACTIVE

    normalized = slugify(raw_status).replace("-", "").lower()
    if normalized in EMPLOYEE_STATUS_ALIASES:
        return EMPLOYEE_STATUS_ALIASES[normalized]

    status = ALLOWED_EMPLOYEE_STATUSES.get(normalized)
    if not status:
        raise ValueError("Status de usuário inválido. Use 'ativo' ou 'inativo'.")
    return status


def _normalize_origem(raw_origem: str | None) -> str:
    if not raw_origem:
        return ""
    normalized = raw_origem.strip().upper()
    if normalized in PhoneLine.Origem.values:
        return normalized
    raise ValueError(
        f"Origem inválida: '{raw_origem}'. "
        f"Valores aceitos: {', '.join(PhoneLine.Origem.values)}."
    )


def _normalize_phone_line_status(raw_status: str | None) -> str:
    if not raw_status:
        return PhoneLine.Status.AVAILABLE

    normalized = slugify(raw_status).replace("-", "").lower()
    if normalized in PHONE_LINE_STATUS_ALIASES:
        return PHONE_LINE_STATUS_ALIASES[normalized]

    status = ALLOWED_PHONE_LINE_STATUSES.get(normalized)
    if not status:
        raise ValueError(
            "Status de linha invalido para upload de simcard. "
            "Use AVAILABLE/ALLOCATED/SUSPENDED/CANCELLED/AQUECENDO/NOVO "
            "ou equivalentes em portugues."
        )
    return status


def _map_phone_line_status_to_sim_status(phone_line_status: str) -> str:
    mapping = {
        PhoneLine.Status.AVAILABLE: SIMcard.Status.AVAILABLE,
        PhoneLine.Status.ALLOCATED: SIMcard.Status.ACTIVE,
        PhoneLine.Status.SUSPENDED: SIMcard.Status.BLOCKED,
        PhoneLine.Status.CANCELLED: SIMcard.Status.CANCELLED,
        PhoneLine.Status.AQUECENDO: SIMcard.Status.ACTIVE,
        PhoneLine.Status.NOVO: SIMcard.Status.AVAILABLE,
    }
    return mapping[phone_line_status]
